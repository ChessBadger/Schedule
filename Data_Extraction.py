import os
import openpyxl
import re


# Specify the file path to be deleted
file_path = "extracted_data.txt"

if os.path.exists(file_path):
    os.remove(file_path)
    print(f"{file_path} has been deleted.")


# Function to search for "Lashaun" in a worksheet and extract links and times
def extract_info(ws):
    days = []
    days_to_numbers = {
        "Sunday": 2,
        "Monday": 6,
        "Tuesday": 10,
        "Wednesday": 14,
        "Thursday": 18,
        "Friday": 22,
        "Saturday": 26
    }
    links = []
    times = []
    second_times = []
    locations = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "LASHAUN" in str(cell.value).upper():
                # Search for a link moving up
                day = None
                # Check the cell above and up to 3 cells above
                for offset in range(-1, -100, -1):
                    if cell.row + offset <= 1:  # Stop if we reach the top of the sheet
                        # Set above_cell to the top cell
                        above_cell = ws.cell(row=1, column=cell.column)
                        for k, v in days_to_numbers.items():
                            if v == above_cell.column:
                                day = k
                        break
                    above_cell = ws.cell(
                        row=cell.row + offset, column=cell.column)
                if day:
                    days.append(day)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "LASHAUN" in str(cell.value).upper():
                # Search for a link moving up
                link = None
                # Check the cell above and up to 3 cells above
                for offset in range(-1, -100, -1):
                    if cell.row + offset <= 1:  # Stop if we reach the top of the sheet
                        break
                    above_cell = ws.cell(
                        row=cell.row + offset, column=cell.column)
                    if re.match(r'https?://\S+', str(above_cell.value)) or "OFFICE" in str(above_cell.value).upper():
                        link = above_cell.value
                        break
                if link:
                    links.append(link)

    for link in links:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and link in str(cell.value):
                    # Search for a time moving up
                    location = None
                    if "OFFICE" in str(cell.value).upper():
                        location = cell.value
                        locations.append(location)
                        break
                    # Check the cell above the link and up to 3 cells above
                    for offset in range(-1, -3, -1):
                        if cell.row + offset <= 1:
                            break
                        above_cell = ws.cell(
                            row=cell.row + offset, column=cell.column)
                        if (r'^[^0-9!@#$%^&*()_=+[\]{};:\'",.<>?/\\|][^]*$', str(above_cell.value)):
                            location = above_cell.value.replace("\n", " ")
                    if location:
                        locations.append(location)
                        row = cell.row

    for link in links:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and link in str(cell.value):
                    # Search for a time moving up
                    time = None
                    if "OFFICE" in str(cell.value).upper():
                        time = "00:00"
                        times.append(time)
                        break
                    # Check the cell above the link and up to 3 cells above
                    for offset in range(-1, -100, -1):
                        if cell.row + offset <= 1:
                            break
                        above_cell = ws.cell(
                            row=cell.row + offset, column=cell.column)
                        if re.match(r'\d{1,2}:\d{2}', str(above_cell.value)):
                            time = above_cell.value
                            break
                    if time:
                        times.append(time)
                        row = cell.row

    for link in links:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and link in str(cell.value):
                    # Search for a second time moving up (within 3 cells) from the time
                    second_time = None
                    # Only move up three cells this time
                    for offset in range(-1, -10, -1):
                        if cell.row + offset <= 1:
                            break
                        above_cell = ws.cell(
                            row=cell.row + offset, column=cell.column)
                        if re.match(r'\d{1,2}:\d{2}', str(above_cell.value)):
                            if "MEET" in str(above_cell.value).upper():
                                second_time = above_cell.value
                                break
                            else:
                                second_time = "NO MEET"
                    if second_time:
                        second_times.append(second_time)
                        row = cell.row

    return days, links, locations, second_times, times


# Check the current working directory for Excel files
cwd = os.getcwd()
excel_files = [file for file in os.listdir(cwd) if file.endswith('.xlsx')]


for excel_file in excel_files:
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        days, links, locations, second_times, times = extract_info(ws)

        # Remove newline characters from each item in the 'locations' list
        for i in range(len(locations)):
            locations[i] = locations[i].replace("\n", " ")

        # After extracting information, store it in a list of tuples
        extracted_info = list(zip(days, links, locations, second_times, times))

        # Sort the list of tuples by day
        sorted_info = sorted(extracted_info, key=lambda x: x[0])

        # Create a dictionary to map days to their order
        day_order = {
            "Sunday": 0,
            "Monday": 1,
            "Tuesday": 2,
            "Wednesday": 3,
            "Thursday": 4,
            "Friday": 5,
            "Saturday": 6
        }

        # Sort the sorted_info list using the day_order dictionary
        sorted_info = sorted(sorted_info, key=lambda x: day_order[x[0]])

        # Iterate through the sorted information and print it
        for info in sorted_info:
            day, link, location, second_time, time = info
            print(day + "\n" + link + "\n" + location +
                  "\n" + second_time + "\n" + time + "\n")


# Create a text file to store the extracted data
output_file = open("extracted_data.txt", "w")

for excel_file in excel_files:
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file)

    # Iterate through all sheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        days, links, locations, second_times, times = extract_info(ws)

        extracted_info = list(zip(days, links, locations, second_times, times))

        sorted_info = sorted(extracted_info, key=lambda x: x[0])

        day_order = {
            "Sunday": 0,
            "Monday": 1,
            "Tuesday": 2,
            "Wednesday": 3,
            "Thursday": 4,
            "Friday": 5,
            "Saturday": 6
        }

        sorted_info = sorted(sorted_info, key=lambda x: day_order[x[0]])

        for info in sorted_info:
            day, link, locations, second_time, time = info
            data_to_write = f"{day}\n{link}\n{locations}\n{second_time}\n{time}\n"

            # Write the data to the output file
            output_file.write(data_to_write)

# Close the output file
output_file.close()
print("Data has been written to extracted_data.txt")
